import tempfile
import unittest
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import Workbook

from app.services.rag_package_builder import (
    RAG_TEXT_EXTRACTION_EXTENSIONS,
    extract_text_for_rag,
    split_spreadsheet_into_chunks,
)


class RagPackageBuilderExcelTests(unittest.TestCase):
    def test_xlsx_extracts_all_sheets_rows_formats_merges_and_formulas(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "업무현황.xlsx"
            self._create_workbook(path)
            self._set_formula_cache(path, "E2", "5000000")

            text = extract_text_for_rag(str(path))

        self.assertIn("[시트: 매출현황]", text)
        self.assertIn("[시트: 참고자료]", text)
        self.assertIn("품목: 노트북", text)
        self.assertIn("비율: 15%", text)
        self.assertIn("금액: ₩1,000,000", text)
        self.assertIn("계산금액: ₩5,000,000", text)
        self.assertIn("병합메모: 공통 메모", text)
        self.assertIn("열 7: 공통 메모", text)
        self.assertIn("미계산: 수식 결과를 확인할 수 없습니다", text)
        self.assertIn("품목: 항목15", text)
        self.assertIn("구분: 별도 시트 값", text)

    def test_spreadsheet_chunks_preserve_complete_row_boundaries(self):
        rows = ["[시트: 매출현황]", "[헤더] 품목 | 금액"]
        rows.extend(
            f"[행 {index}] 품목: 항목{index} | 금액: ₩{index * 1000:,}"
            for index in range(2, 45)
        )
        chunks = split_spreadsheet_into_chunks("\n".join(rows), chunk_size=180)

        self.assertGreater(len(chunks), 1)
        self.assertTrue(all(chunk.startswith("[시트: 매출현황]") for chunk in chunks))
        for row in rows[1:]:
            self.assertEqual(sum(row in chunk for chunk in chunks), 1)

    def test_xlsx_and_xls_are_in_embedding_whitelist(self):
        self.assertIn(".xlsx", RAG_TEXT_EXTRACTION_EXTENSIONS)
        self.assertIn(".xls", RAG_TEXT_EXTRACTION_EXTENSIONS)

    @staticmethod
    def _create_workbook(path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "매출현황"
        sheet.append(["품목", "수량", "비율", "금액", "계산금액", "병합메모", "", "미계산"])
        sheet.append(["노트북", 5, 0.15, 1_000_000, "=B2*D2", "공통 메모", "", "=1+1"])
        sheet.merge_cells("F2:G2")
        sheet["C2"].number_format = "0%"
        sheet["D2"].number_format = "₩#,##0"
        sheet["E2"].number_format = "₩#,##0"
        for index in range(3, 17):
            sheet.append([f"항목{index - 1}", index, index / 100, index * 10_000])
            sheet.cell(index, 3).number_format = "0%"
            sheet.cell(index, 4).number_format = "₩#,##0"

        second_sheet = workbook.create_sheet("참고자료")
        second_sheet.append(["구분", "내용"])
        second_sheet.append(["별도 시트 값", "두 번째 시트도 검색 대상"])
        workbook.save(path)

    @staticmethod
    def _set_formula_cache(path: Path, coordinate: str, value: str) -> None:
        with zipfile.ZipFile(path) as archive:
            entries = {name: archive.read(name) for name in archive.namelist()}

        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        root = ET.fromstring(entries["xl/worksheets/sheet1.xml"])
        cell = root.find(f".//{{{namespace}}}c[@r='{coordinate}']")
        if cell is None:
            raise AssertionError(f"Formula cell not found: {coordinate}")
        value_node = cell.find(f"{{{namespace}}}v")
        if value_node is None:
            value_node = ET.SubElement(cell, f"{{{namespace}}}v")
        value_node.text = value
        entries["xl/worksheets/sheet1.xml"] = ET.tostring(
            root,
            encoding="utf-8",
            xml_declaration=True,
        )

        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, content in entries.items():
                archive.writestr(name, content)


if __name__ == "__main__":
    unittest.main()
